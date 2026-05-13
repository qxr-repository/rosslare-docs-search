#!/usr/bin/env python3
import json
import re
import html as html_escape
from pathlib import Path
from datetime import datetime

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except:
    HAS_DOCX = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except:
    HAS_OPENPYXL = False

class EnhancedContentIndexer:
    def __init__(self, root_path):
        self.root_path = Path(root_path)
        self.index = []
        
    def extract_docx_formatted(self, file_path):
        """Extract DOCX with formatting as HTML"""
        if not HAS_DOCX:
            return None, None
        try:
            doc = DocxDocument(file_path)
            plain_text = []
            html_content = ['<div style="font-family: Arial; line-height: 1.6;">']
            
            for para in doc.paragraphs:
                if para.text.strip():
                    plain_text.append(para.text)
                    
                    # Extract formatting
                    html_para = '<p style="margin: 8px 0;">'
                    for run in para.runs:
                        text = html_escape.escape(run.text)
                        style = ''
                        if run.bold:
                            style += 'font-weight: bold;'
                        if run.italic:
                            style += 'font-style: italic;'
                        if run.underline:
                            style += 'text-decoration: underline;'
                        
                        if style:
                            html_para += f'<span style="{style}">{text}</span>'
                        else:
                            html_para += text
                    html_para += '</p>'
                    html_content.append(html_para)
            
            html_content.append('</div>')
            
            plain = ' '.join(plain_text) if plain_text else None
            formatted = ''.join(html_content) if len(html_content) > 2 else None
            
            return plain, formatted
        except:
            return None, None
    
    def extract_excel_formatted(self, file_path):
        """Extract XLSX as HTML table"""
        if not HAS_OPENPYXL:
            return None, None
        try:
            workbook = load_workbook(file_path)
            plain_text = []
            html_content = ['<div style="overflow-x: auto; font-family: Arial;">']
            
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                
                # Add sheet name
                html_content.append(f'<h4 style="margin: 10px 0 5px 0; color: #667eea;">{html_escape.escape(sheet_name)}</h4>')
                html_content.append('<table style="border-collapse: collapse; border: 1px solid #ddd; font-size: 12px;">')
                
                for row in ws.iter_rows(values_only=True):
                    # Collect plain text
                    for cell in row:
                        if cell is not None:
                            plain_text.append(str(cell))
                    
                    # Build HTML row
                    html_content.append('<tr>')
                    for cell in row:
                        cell_text = str(cell) if cell is not None else ''
                        cell_text = html_escape.escape(cell_text)
                        html_content.append(f'<td style="border: 1px solid #ddd; padding: 6px; background: #f9f9f9;">{cell_text}</td>')
                    html_content.append('</tr>')
                
                html_content.append('</table>')
                html_content.append('<br>')
            
            html_content.append('</div>')
            
            plain = ' '.join(plain_text) if plain_text else None
            formatted = ''.join(html_content) if len(html_content) > 2 else None
            
            return plain, formatted
        except:
            return None, None
    
    def extract_text_file(self, file_path):
        """Extract TXT/MD/HTML with basic formatting"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Simple formatting for plain text
            formatted = f'<pre style="font-family: monospace; line-height: 1.4; white-space: pre-wrap; word-wrap: break-word;">{html_escape.escape(content)}</pre>'
            
            return content, formatted
        except:
            return None, None
    
    def extract_content(self, file_path):
        """Extract content from supported types - returns (plain_text, formatted_html)"""
        suffix = file_path.suffix.lower()
        
        if suffix == '.docx':
            return self.extract_docx_formatted(file_path)
        elif suffix in ['.xls', '.xlsx']:
            return self.extract_excel_formatted(file_path)
        elif suffix in ['.txt', '.md', '.html']:
            return self.extract_text_file(file_path)
        
        return None, None
    
    def get_keywords(self, text):
        """Fast keyword extraction"""
        if not text:
            return []
        
        words = re.findall(r'\b\w{3,}\b', text.lower())
        freq = {}
        for w in words:
            freq[w] = freq.get(w, 0) + 1
        
        top = sorted(freq.items(), key=lambda x: x[1], reverse=True)[:40]
        return [w for w, _ in top]
    
    def index_file(self, file_path):
        """Index a file with formatted content"""
        try:
            rel_path = file_path.relative_to(self.root_path)
            file_size = file_path.stat().st_size
            
            # Extract content
            plain_content, formatted_content = self.extract_content(file_path)
            
            if plain_content:
                # Limit plain text
                plain_content = ' '.join(plain_content.split())[:40000]
                
                # Get keywords
                filename_words = re.findall(r'\b\w+\b', file_path.stem)
                keywords = self.get_keywords(plain_content + ' ' + ' '.join(filename_words))
                
                snippet = plain_content[:300] + ("..." if len(plain_content) > 300 else "")
                
                entry = {
                    'path': str(rel_path).replace('\\', '/'),
                    'filename': file_path.name,
                    'file_type': file_path.suffix.lower(),
                    'size_kb': round(file_size / 1024, 1),
                    'snippet': snippet,
                    'keywords': keywords,
                    'content': plain_content,
                    'formatted_content': formatted_content or ''  # Store formatted HTML
                }
                
                self.index.append(entry)
                return True
        except Exception as e:
            print(f"Error indexing {file_path}: {e}")
        
        return False
    
    def build_index(self):
        """Build index for text-based files"""
        print("Building enhanced content index with formatted previews...\n")
        
        all_files = list(self.root_path.rglob('*'))
        text_files = [f for f in all_files if f.is_file() and f.suffix in 
                     ['.docx', '.txt', '.md', '.html', '.xls', '.xlsx']]
        
        print(f"Found {len(text_files)} text-based files\n")
        
        for i, f in enumerate(sorted(text_files)):
            if i % 10 == 0:
                print(f"  {i}/{len(text_files)}...")
            self.index_file(f)
        
        print(f"\n✓ Indexed {len(self.index)} files with formatted content")
        return self.index
    
    def save_index(self, output_dir):
        """Save index"""
        output_file = Path(output_dir) / 'content_index.json'
        
        total_size = sum(len(f.get('content', '')) + len(f.get('formatted_content', '')) for f in self.index)
        
        index_data = {
            'created': datetime.now().isoformat(),
            'total_files': len(self.index),
            'total_content_mb': round(total_size / 1024 / 1024, 2),
            'files': self.index
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(index_data, f, indent=2, ensure_ascii=False)
        
        file_size_mb = output_file.stat().st_size / 1024 / 1024
        print(f"✓ Index saved: {output_file}")
        print(f"  Size: {file_size_mb:.1f} MB")
        return output_file

if __name__ == '__main__':
    indexer = EnhancedContentIndexer('/sessions/gifted-sharp-feynman/mnt/Rosslare Documentation/')
    indexer.build_index()
    indexer.save_index('/sessions/gifted-sharp-feynman/mnt/outputs')
